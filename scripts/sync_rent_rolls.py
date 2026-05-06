#!/usr/bin/env python3
"""
sync_rent_rolls.py — Pull Yardi rent-roll xlsx exports from Dropbox and upsert
into Supabase `rent_roll` table. Same pattern as sync_actuals.py.

Usage:
    python3 scripts/sync_rent_rolls.py                # dry-run
    python3 scripts/sync_rent_rolls.py --commit       # write to Supabase
    python3 scripts/sync_rent_rolls.py --property p0000003
    python3 scripts/sync_rent_rolls.py --file path/to/RentRoll.xlsx
    python3 scripts/sync_rent_rolls.py --verbose

Looks in each property folder for files matching common rent-roll naming:
    - "Rent Roll*.xlsx"
    - "rent_roll*.xlsx"
    - "RR*.xlsx"
    - "*rentroll*.xlsx" / "*rent-roll*.xlsx"
Prefers files inside "3 - Operations", "2 - Leasing_Marketing", or
"4 - Accounting/Monthly Financials Reports".

NOTE: Yardi rent roll column names vary by tenant/property. This parser uses
fuzzy header matching. When new formats appear, extend COLUMN_ALIASES.
"""

from __future__ import annotations

import argparse
import glob
import os
import re
import sys
import time
from datetime import date, datetime
from pathlib import Path
from typing import Any

try:
    import openpyxl
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl requests")


# ─────────────────────────────────────────────────────────────────────────
# Paths (mirror sync_actuals.py)
# ─────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
CONFIG_JS = PROJECT_DIR / "config.js"

DROPBOX_ROOTS = [
    Path.home() / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "Dropbox-FirstMileCapital",
    Path.home() / "First Mile Dropbox" / "Morris Zeitouni" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox",
]
PROPMGMT_DIRNAME = "2.1 FMC Property Management"

# Preferred subfolders to search (in order)
PREFERRED_SUBDIRS = [
    "3 - Operations",
    "2 - Leasing_Marketing",
    "2 - Leasing",
    "4 - Accounting",
]

# Yardi property-code → Supabase property_id (same mapping as sync_actuals.py)
PROPERTY_MAP: dict[str, dict[str, str]] = {
    "p0000003": {"id": "recQX1kpeJKqIzvkU", "name": "Paramus Plaza"},
    "p0000004": {"id": "recUUsUChvL3yQ96g", "name": "340 Mount Kemble"},
    "p0000005": {"id": "recqfxJfdqCXCLOuD", "name": "61 S Paramus"},
    "p0000006": {"id": "recxF4R64gbb5Sowj", "name": "575 Broadway"},
    "p0000007": {"id": "recF3zFKbY4wJ4P40", "name": "1700 East Putnam"},
}
FOLDER_TO_PROPERTY: dict[str, str] = {
    "paramus plaza": "recQX1kpeJKqIzvkU",
    "340 mt kemble morristown": "recUUsUChvL3yQ96g",
    "340 mount kemble": "recUUsUChvL3yQ96g",
    "61 s paramus": "recqfxJfdqCXCLOuD",
    "575 broadway": "recxF4R64gbb5Sowj",
    "1700 east putnam greenwich": "recF3zFKbY4wJ4P40",
    "1700 east putnam": "recF3zFKbY4wJ4P40",
}

FOLDER_SKIPLIST = [
    "0. overall property management",
    "41 flatbush",
    "red bank - 1 river centre",
    "red bank",
]


# ─────────────────────────────────────────────────────────────────────────
# Column alias map for fuzzy header matching.
# Key = canonical field name, value = list of header-text variants to recognize.
# All comparisons are case-insensitive + whitespace-normalized.
# ─────────────────────────────────────────────────────────────────────────
COLUMN_ALIASES: dict[str, list[str]] = {
    "tenant_name":        ["tenant", "tenant name", "resident", "lessee"],
    "suite":              ["suite", "unit", "space", "suite #", "unit #"],
    "status":             ["status", "lease status", "tenant status"],
    "sf":                 ["sf", "sq ft", "square feet", "rsf", "leased sf", "rentable sf", "area"],
    "lease_start":        ["lease start", "lease from", "commencement", "start date", "lease commencement", "move in", "move-in"],
    "lease_end":          ["lease end", "lease to", "expiration", "end date", "lease expiration", "term end"],
    "move_in_date":       ["move in date", "move-in date", "rent commencement"],
    "move_out_date":      ["move out date", "move-out date", "vacate date"],
    "monthly_rent":       ["monthly rent", "current rent", "current monthly rent", "base rent / mo", "monthly base rent", "base rent (monthly)"],
    "annual_rent":        ["annual rent", "current annual rent", "annual base rent", "base rent (annual)", "yearly rent"],
    "rent_per_sf":        ["rent/sf", "rent psf", "$/sf", "rate/sf", "rate psf", "base rent psf"],
    "escalation_pct":     ["escalation", "escalation %", "annual escalation", "bump", "esc %"],
    "next_escalation_date": ["next increase", "next escalation", "next bump", "escalation date"],
    "free_rent_months":   ["free rent", "free rent months", "rent abatement"],
    "cam_reimbursement":  ["cam", "cam reimbursement", "cam recovery", "cam method"],
    "re_tax_reimbursement": ["tax reimbursement", "re tax", "real estate tax recovery"],
    "insurance_reimbursement": ["insurance reimbursement", "insurance recovery"],
    "notes":              ["notes", "comments", "remarks"],
}


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────
def norm_header(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip().lower()
    s = re.sub(r"[_\-]+", " ", s)
    s = re.sub(r"\s+", " ", s)
    return s


def build_header_index(header_row: tuple) -> dict[str, int]:
    """Map each canonical field name to the column index that matches best."""
    out: dict[str, int] = {}
    normed = [norm_header(c) for c in header_row]
    for field, aliases in COLUMN_ALIASES.items():
        for alias in aliases:
            aliasN = norm_header(alias)
            for i, cell in enumerate(normed):
                if cell == aliasN or cell.startswith(aliasN):
                    out.setdefault(field, i)
                    break
            if field in out:
                break
    return out


def to_date(v: Any) -> str | None:
    if v is None or v == "":
        return None
    if isinstance(v, (datetime, date)):
        return v.isoformat()[:10]
    s = str(v).strip()
    # Try a few common formats
    for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y", "%d-%b-%Y", "%b %d, %Y", "%m-%d-%Y"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def to_num(v: Any) -> float | None:
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip().replace("$", "").replace(",", "").replace("%", "")
    if s in ("", "-", "—", "N/A"):
        return None
    try:
        return float(s)
    except ValueError:
        return None


def find_dropbox_root() -> Path | None:
    for p in DROPBOX_ROOTS:
        if p.exists() and p.is_dir():
            return p
    return None


def find_rent_roll_file(prop_folder: Path) -> Path | None:
    """Search the property folder for the latest rent roll xlsx."""
    patterns = [
        "Rent Roll*.xlsx", "rent roll*.xlsx",
        "RentRoll*.xlsx", "rent_roll*.xlsx",
        "RR*.xlsx",
        "*rent roll*.xlsx", "*rentroll*.xlsx",
        "*Rent Roll*.xlsx", "*RentRoll*.xlsx",
    ]
    # Prefer certain subdirs first
    candidates: list[Path] = []
    for sub in PREFERRED_SUBDIRS:
        sub_path = prop_folder / sub
        if not sub_path.exists():
            continue
        for pat in patterns:
            for p in glob.glob(str(sub_path / pat)):
                candidates.append(Path(p))
            for p in glob.glob(str(sub_path / "**" / pat), recursive=True):
                candidates.append(Path(p))
    # Fallback — search the whole property folder
    if not candidates:
        for pat in patterns:
            for p in glob.glob(str(prop_folder / "**" / pat), recursive=True):
                candidates.append(Path(p))
    # Dedupe + filter Excel lock files
    seen = set()
    unique: list[Path] = []
    for p in candidates:
        if p.name.startswith("~$"):
            continue
        key = str(p)
        if key in seen:
            continue
        seen.add(key)
        unique.append(p)
    if not unique:
        return None
    unique.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return unique[0]


def materialize_dropbox_file(path: Path, timeout: int = 300) -> bool:
    try:
        if path.stat().st_size > 0:
            return True
    except OSError:
        return False
    import subprocess
    try:
        subprocess.run(["/bin/cat", str(path)], stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL, timeout=timeout, check=False)
    except Exception:
        pass
    try:
        return path.stat().st_size > 0
    except OSError:
        return False


def read_supabase_config() -> tuple[str, str]:
    env_url = os.environ.get("SUPABASE_URL")
    env_key = os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_SERVICE_KEY")
    if env_url and env_key:
        return env_url.rstrip("/"), env_key
    if not CONFIG_JS.exists():
        sys.exit("config.js not found. Set SUPABASE_URL + SUPABASE_KEY env vars.")
    text = CONFIG_JS.read_text()
    url_m = re.search(r"SUPABASE_URL\s*=\s*['\"]([^'\"]+)['\"]", text)
    key_m = re.search(r"SUPABASE_KEY\s*=\s*['\"]([^'\"]+)['\"]", text)
    if not (url_m and key_m):
        sys.exit("Could not parse SUPABASE_URL / SUPABASE_KEY from config.js")
    return url_m.group(1).rstrip("/"), key_m.group(1)


# ─────────────────────────────────────────────────────────────────────────
# Parser — Yardi "charge-line" format
# Each tenant is a block with one row per charge code (rent-com, CAM-EST, …)
# and one row per unit. Header row contains "Charge Code" + "Lease From" +
# "Monthly Amount". Tenant name lives in the "Lease" column on the first row
# of each block; subsequent rows leave that column blank. Block ends with a
# row whose first column reads "Total".
# ─────────────────────────────────────────────────────────────────────────
def parse_yardi_charge_format(path: Path, verbose: bool = False):
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Locate "as of" date (e.g. row 2 = "As of Date: 2/28/2026")
    as_of_d: date | None = None
    for r in all_rows[:8]:
        if not r:
            continue
        for c in r:
            if c is None:
                continue
            s = str(c).strip()
            m = re.search(r"as of(?:\s*date)?:?\s*(\d{1,2}/\d{1,2}/\d{2,4})",
                          s, re.IGNORECASE)
            if m:
                for fmt in ("%m/%d/%Y", "%m/%d/%y"):
                    try:
                        as_of_d = datetime.strptime(m.group(1), fmt).date()
                        break
                    except ValueError:
                        continue
            if as_of_d:
                break
        if as_of_d:
            break

    # Locate header row
    header_idx = None
    for i, row in enumerate(all_rows[:15]):
        if not row:
            continue
        normed = " | ".join(norm_header(c) for c in row)
        if "charge code" in normed and "lease from" in normed and "monthly amount" in normed:
            header_idx = i
            break
    if header_idx is None:
        return None  # not in this format

    headers = [norm_header(c) for c in all_rows[header_idx]]

    def find_col(name: str, occurrence: int = 0):
        n = norm_header(name)
        found = 0
        for i, h in enumerate(headers):
            if h == n:
                if found == occurrence:
                    return i
                found += 1
        return None

    col_units_a    = find_col("unit(s)", 0)
    col_lease      = find_col("lease", 0)
    col_lease_from = find_col("lease from")
    col_lease_to   = find_col("lease to")
    col_units_i    = find_col("unit(s)", 1) or col_units_a
    col_area       = find_col("area")
    col_charge     = find_col("charge code")
    col_chg_from   = find_col("charge from")
    col_chg_to     = find_col("charge to")
    col_monthly    = find_col("monthly amount")
    col_annualized = find_col("annualized gross amount")

    def cell(row, idx):
        if idx is None or idx >= len(row):
            return None
        return row[idx]

    def to_dt(v: Any):
        if v is None or v == "":
            return None
        if isinstance(v, datetime):
            return v.date()
        if isinstance(v, date):
            return v
        for fmt in ("%Y-%m-%d", "%m/%d/%Y", "%m/%d/%y"):
            try:
                return datetime.strptime(str(v).strip()[:10], fmt).date()
            except ValueError:
                continue
        return None

    def is_active(cf, ct) -> bool:
        if not as_of_d:
            return True
        cfd = to_dt(cf)
        ctd = to_dt(ct)
        if cfd and as_of_d < cfd:
            return False
        if ctd and as_of_d > ctd:
            return False
        return True

    suites: dict[tuple[str, str], dict[str, Any]] = {}
    current_tenant: str | None = None
    current_ls: str | None = None
    current_le: str | None = None

    for row in all_rows[header_idx + 1:]:
        if not row or all(c is None or str(c).strip() == "" for c in row):
            continue
        a = cell(row, col_units_a)
        b = cell(row, col_lease)
        a_str = str(a).strip() if a is not None else ""
        b_str = str(b).strip() if b is not None else ""

        if a_str.lower() in ("total", "subtotal", "grand total"):
            current_tenant = None
            continue

        if b_str:
            current_tenant = b_str
            current_ls = to_date(cell(row, col_lease_from))
            current_le = to_date(cell(row, col_lease_to))

        if not current_tenant:
            continue

        unit_raw = cell(row, col_units_i)
        unit_str = str(unit_raw).strip() if unit_raw is not None else ""
        if not unit_str:
            continue

        area = to_num(cell(row, col_area))
        charge = cell(row, col_charge)
        charge_str = str(charge).strip() if charge is not None else ""
        if not charge_str:
            continue

        cf = cell(row, col_chg_from)
        ct = cell(row, col_chg_to)
        monthly = to_num(cell(row, col_monthly)) or 0.0
        annual = to_num(cell(row, col_annualized)) or (monthly * 12)
        active = is_active(cf, ct)

        key = (current_tenant, unit_str)
        s = suites.setdefault(key, {
            "tenant_name": current_tenant,
            "suite": unit_str,
            "status": "Current",
            "sf": area,
            "lease_start": current_ls,
            "lease_end": current_le,
            "monthly_rent": 0.0,
            "annual_rent": 0.0,
            "cam_reimbursement": None,
            "re_tax_reimbursement": None,
            "insurance_reimbursement": None,
            "_other": [],
        })
        if not s.get("sf") and area:
            s["sf"] = area
        if not s.get("lease_start") and current_ls:
            s["lease_start"] = current_ls
        if not s.get("lease_end") and current_le:
            s["lease_end"] = current_le

        if not active:
            continue
        cc = charge_str.lower()
        if any(k in cc for k in ("rent-com", "rent com", "base rent")) or (
            cc.startswith("rent") and not any(k in cc for k in ("free", "abate", "concession"))
        ):
            s["monthly_rent"] += monthly
            s["annual_rent"] += annual
        elif "cam" in cc or cc.startswith("oe") or "operating expense" in cc:
            lbl = f"{charge_str}: ${monthly:,.0f}/mo"
            s["cam_reimbursement"] = (s["cam_reimbursement"] + "; " if s["cam_reimbursement"] else "") + lbl
        elif "tax" in cc or cc == "ret":
            lbl = f"{charge_str}: ${monthly:,.0f}/mo"
            s["re_tax_reimbursement"] = (s["re_tax_reimbursement"] + "; " if s["re_tax_reimbursement"] else "") + lbl
        elif "ins" in cc:
            lbl = f"{charge_str}: ${monthly:,.0f}/mo"
            s["insurance_reimbursement"] = (s["insurance_reimbursement"] + "; " if s["insurance_reimbursement"] else "") + lbl
        else:
            s["_other"].append(f"{charge_str}: ${monthly:,.0f}/mo")

    rows_out: list[dict[str, Any]] = []
    for (tenant, unit), s in suites.items():
        status = "Current"
        if as_of_d and s.get("lease_end"):
            le = to_dt(s["lease_end"])
            if le and as_of_d > le:
                status = "Expired"
        if as_of_d and s.get("lease_start"):
            ls = to_dt(s["lease_start"])
            if ls and as_of_d < ls:
                status = "Pending"
        rec = {
            "tenant_name": s["tenant_name"],
            "suite": s["suite"],
            "status": status,
            "sf": s.get("sf"),
            "lease_start": s.get("lease_start"),
            "lease_end": s.get("lease_end"),
            "monthly_rent": round(s["monthly_rent"], 2) if s["monthly_rent"] else None,
            "annual_rent": round(s["annual_rent"], 2) if s["annual_rent"] else None,
            "cam_reimbursement": s.get("cam_reimbursement"),
            "re_tax_reimbursement": s.get("re_tax_reimbursement"),
            "insurance_reimbursement": s.get("insurance_reimbursement"),
        }
        if rec["annual_rent"] and rec.get("sf"):
            try:
                rec["rent_per_sf"] = round(rec["annual_rent"] / rec["sf"], 4)
            except (ZeroDivisionError, TypeError):
                pass
        if s.get("_other"):
            rec["notes"] = "; ".join(s["_other"])
        rows_out.append(rec)

    if verbose:
        print(f"    yardi-charge-format: as_of={as_of_d}, suites={len(rows_out)}")

    return {
        "rows": rows_out,
        "source_file": str(path),
        "header_columns": ["yardi_charge_format"],
        "as_of_date": as_of_d.isoformat() if as_of_d else None,
    }


# ─────────────────────────────────────────────────────────────────────────
# Parser — generic / legacy table format (one row per tenant)
# ─────────────────────────────────────────────────────────────────────────
def parse_rent_roll(path: Path, verbose: bool = False) -> dict[str, Any]:
    # Try the Yardi charge-line format first
    try:
        yardi = parse_yardi_charge_format(path, verbose=verbose)
        if yardi is not None and yardi.get("rows"):
            return yardi
    except Exception as e:
        if verbose:
            print(f"    yardi parser exception: {e}")

    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Find the header row — look for the row containing "tenant" or "unit" or "suite" in the first ~20 rows
    header_idx = None
    for i, row in enumerate(all_rows[:25]):
        if not row:
            continue
        normed = [norm_header(c) for c in row]
        has_tenant = any(("tenant" in c) or c == "lessee" or ("resident" in c) for c in normed)
        has_anchor = any(c in {"suite", "unit", "sf", "sq ft"} or c.startswith("sq ft") for c in normed)
        if has_tenant and has_anchor:
            header_idx = i
            break

    if header_idx is None:
        raise ValueError(f"Could not locate header row in {path.name}. Inspect manually.")

    header = all_rows[header_idx]
    col_idx = build_header_index(header)

    if "tenant_name" not in col_idx:
        raise ValueError(f"Header row in {path.name} has no 'tenant' column.")

    out_rows: list[dict[str, Any]] = []
    for row in all_rows[header_idx + 1 :]:
        if not row:
            continue
        # Skip totals/subtotals (typically in first column)
        first = str(row[0]).strip().lower() if row[0] is not None else ""
        if first.startswith(("total", "subtotal", "grand total", "all tenants")):
            continue

        rec: dict[str, Any] = {}
        for field, idx in col_idx.items():
            val = row[idx] if idx < len(row) else None
            if val is None or str(val).strip() == "":
                continue
            if field in ("lease_start", "lease_end", "move_in_date", "move_out_date", "next_escalation_date"):
                rec[field] = to_date(val)
            elif field in ("sf", "monthly_rent", "annual_rent", "rent_per_sf", "escalation_pct",
                           "free_rent_months", "cam_rate_psf", "re_tax_rate_psf", "insurance_rate_psf"):
                rec[field] = to_num(val)
            else:
                rec[field] = str(val).strip()

        # Require at least a tenant name or a status of 'Vacant'
        if not rec.get("tenant_name") and not rec.get("status"):
            continue

        # Derive annual_rent / rent_per_sf if missing
        if not rec.get("annual_rent") and rec.get("monthly_rent"):
            rec["annual_rent"] = round(rec["monthly_rent"] * 12, 2)
        if not rec.get("rent_per_sf") and rec.get("annual_rent") and rec.get("sf"):
            try:
                rec["rent_per_sf"] = round(rec["annual_rent"] / rec["sf"], 4)
            except ZeroDivisionError:
                pass

        out_rows.append(rec)

    # Generic parser doesn't have an explicit "As of" header, so fall back
    # to the file's modified-time as the snapshot date.
    try:
        mtime = datetime.fromtimestamp(path.stat().st_mtime).date()
        as_of = mtime.isoformat()
    except OSError:
        as_of = None

    return {
        "rows": out_rows,
        "source_file": str(path),
        "header_columns": list(col_idx.keys()),
        "as_of_date": as_of,
    }


def infer_property_id_from_folder(folder_name: str) -> str | None:
    key = folder_name.lower().strip()
    if key in FOLDER_TO_PROPERTY:
        return FOLDER_TO_PROPERTY[key]
    for k, v in FOLDER_TO_PROPERTY.items():
        if k in key or key in k:
            return v
    return None


# ─────────────────────────────────────────────────────────────────────────
# Supabase writer
# ─────────────────────────────────────────────────────────────────────────
def supabase_upsert(url: str, key: str, property_id: str,
                    rows: list[dict[str, Any]], source_file: str,
                    as_of_date: str | None = None) -> int:
    endpoint = f"{url}/rest/v1/rent_roll"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }
    # Delete existing
    del_resp = requests.delete(
        endpoint, headers=headers,
        params={"property_id": f"eq.{property_id}"}, timeout=60,
    )
    if del_resp.status_code not in (200, 204):
        raise RuntimeError(f"DELETE failed ({del_resp.status_code}): {del_resp.text[:400]}")

    if not rows:
        return 0

    # Normalize: PostgREST requires every row in a batch to have identical keys.
    # Build the union of keys across all rows, then fill missing with None.
    all_keys: set[str] = set()
    for r in rows:
        all_keys.update(r.keys())
    all_keys.update({"property_id", "source_file", "as_of_date"})

    payload = []
    for r in rows:
        rec = {k: r.get(k) for k in all_keys}
        rec["property_id"] = property_id
        rec["source_file"] = source_file
        if as_of_date is not None:
            rec["as_of_date"] = as_of_date
        payload.append(rec)

    # Insert in chunks
    chunk = 200
    inserted = 0
    for i in range(0, len(payload), chunk):
        part = payload[i : i + chunk]
        resp = requests.post(endpoint, headers=headers, json=part, timeout=120)
        if resp.status_code not in (200, 201, 204):
            raise RuntimeError(f"POST failed ({resp.status_code}): {resp.text[:400]}")
        inserted += len(part)
    return inserted


# ─────────────────────────────────────────────────────────────────────────
# Discovery
# ─────────────────────────────────────────────────────────────────────────
def discover_rent_rolls(explicit_file: str | None,
                        only_property: str | None) -> dict[str, Path]:
    """Return { folder_name: rent_roll_path }."""
    if explicit_file:
        p = Path(explicit_file).expanduser().resolve()
        if not p.exists():
            sys.exit(f"File not found: {p}")
        return {p.name: p}

    root = find_dropbox_root()
    if root is None:
        print("⚠ Dropbox root not found. Use --file to point at a specific rent roll xlsx.",
              file=sys.stderr)
        return {}

    propmgmt = root / PROPMGMT_DIRNAME
    if not propmgmt.exists():
        print(f"⚠ '{PROPMGMT_DIRNAME}' not found under {root}", file=sys.stderr)
        return {}

    print(f"✓ Scanning {propmgmt}")
    out: dict[str, Path] = {}
    for child in sorted(propmgmt.iterdir()):
        if not child.is_dir():
            continue
        if any(skip in child.name.lower() for skip in FOLDER_SKIPLIST):
            print(f"  ⏭  {child.name}: skiplisted")
            continue
        if only_property and only_property.lower() not in child.name.lower():
            continue
        rr = find_rent_roll_file(child)
        if not rr:
            print(f"  ⏭  {child.name}: no rent roll file found")
            continue
        print(f"  ✓ {child.name}: {rr.name}")
        out[child.name] = rr
    return out


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--commit", action="store_true", help="Actually write to Supabase.")
    ap.add_argument("--property", help="Folder-name substring filter.")
    ap.add_argument("--file", help="Process a single xlsx and stop.")
    ap.add_argument("--verbose", "-v", action="store_true")
    args = ap.parse_args()

    found = discover_rent_rolls(args.file, args.property)
    if not found:
        print("No rent roll files found. Check Dropbox path, or use --file to point at one.")
        return 1

    supabase_url = supabase_key = None
    if args.commit:
        supabase_url, supabase_key = read_supabase_config()

    total_inserted = 0
    for folder_name, path in found.items():
        print(f"\n━━━ {folder_name} ━━━")
        print(f"  file: {path.name}")

        # Materialize if needed
        try:
            sz = path.stat().st_size
        except OSError:
            sz = 0
        if sz == 0:
            print("  … pulling down online-only file from Dropbox …")
            if not materialize_dropbox_file(path):
                print("  ✗ Dropbox download failed (timeout)")
                continue
            print(f"  ✓ materialized ({path.stat().st_size:,} bytes)")

        try:
            parsed = parse_rent_roll(path, verbose=args.verbose)
        except Exception as e:
            print(f"  ✗ parse error: {e}")
            continue

        rows = parsed["rows"]
        as_of_date = parsed.get("as_of_date")
        print(f"  parsed    : {len(rows)} tenant rows")
        print(f"  columns   : {', '.join(parsed['header_columns'])}")
        if as_of_date:
            print(f"  as of     : {as_of_date}")

        if args.verbose:
            for r in rows[:20]:
                print(f"    · {r.get('suite','—'):>6}  {r.get('tenant_name','')[:30]:30s}"
                      f"  {r.get('sf','')!s:>8} SF"
                      f"  rent={r.get('monthly_rent','')!s}"
                      f"  end={r.get('lease_end','')}")

        # Infer property_id from folder name
        supa_id = infer_property_id_from_folder(folder_name)
        if not supa_id:
            print(f"  ✗ Could not map folder '{folder_name}' to a property_id. "
                  f"Add to FOLDER_TO_PROPERTY in this script.")
            continue
        print(f"  property  : {folder_name} → {supa_id}")

        if args.commit:
            n = supabase_upsert(supabase_url, supabase_key, supa_id, rows,
                                str(path), as_of_date=as_of_date)
            total_inserted += n
            print(f"  ✓ upserted {n} rows")
        else:
            print(f"  (dry-run) would upsert {len(rows)} rows")
            total_inserted += len(rows)

    print("\n━━━ summary ━━━")
    print(f"  properties processed : {len(found)}")
    print(f"  total rows           : {total_inserted}")
    print(f"  mode                 : {'COMMIT' if args.commit else 'dry-run'}")
    if not args.commit:
        print("\nRe-run with --commit to actually write.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
