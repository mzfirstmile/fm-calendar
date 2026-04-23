#!/usr/bin/env python3
"""
sync_actuals.py — Pull monthly Yardi 12-month income statements from Dropbox
and upsert actuals into Supabase `actuals_line_items`.

Usage:
    python3 scripts/sync_actuals.py                 # dry-run (shows what would change)
    python3 scripts/sync_actuals.py --commit        # write to Supabase
    python3 scripts/sync_actuals.py --property p0000003  # single property only
    python3 scripts/sync_actuals.py --year 2026     # force year (default: parsed from file)
    python3 scripts/sync_actuals.py --file path/to/file.xlsx  # process one file only

Expects Yardi files at:
    ~/First Mile Dropbox/.../2.1 FMC Property Management/<Property>/4 - Accounting/A - Month Quarter Financials/

Filename pattern: 12Monthsincomestatement_pXXXXXXX_YYYYMMDD_HHMMSS.xlsx
(newest file per property is used)

Account-name → GL-code mapping is derived from data/gl-accounts.json with a
small override table for known aliases (e.g. "Property Tax" → 6512).
"""

from __future__ import annotations

import argparse
import glob
import json
import os
import re
import sys
import time
from collections import defaultdict
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl requests")

try:
    import requests
except ImportError:
    sys.exit("Missing dependency: pip install openpyxl requests")


# ─────────────────────────────────────────────────────────────────────────
# Paths
# ─────────────────────────────────────────────────────────────────────────
SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"
CONFIG_JS = PROJECT_DIR / "config.js"

# Candidate Dropbox roots (first existing wins)
DROPBOX_ROOTS = [
    Path.home() / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "Dropbox-FirstMileCapital",
    Path.home() / "First Mile Dropbox" / "Morris Zeitouni" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "Dropbox" / "First Mile Prop Dropbox",
    Path.home() / "First Mile Dropbox",
]
PROPMGMT_DIRNAME = "2.1 FMC Property Management"
ACCOUNTING_DIRNAME = "4 - Accounting"
# Accountant publishes here. Capitalization/spaces vary; we match case-insensitively.
MONTH_FINANCIALS_DIRNAMES = [
    "A - Month Quarter Financials",
    "A - Month-Quarter Financials",
    "A - month quarter financials",
    "A. Month Quarter Financials",
]


# ─────────────────────────────────────────────────────────────────────────
# Property code (Yardi) ↔ Supabase property_id
# ─────────────────────────────────────────────────────────────────────────
# Derived from files in data/2025 Actuals/ and FB_PROP_META in index.html.
PROPERTY_MAP: dict[str, dict[str, str]] = {
    "p0000003": {"id": "recQX1kpeJKqIzvkU", "name": "Paramus Plaza"},
    "p0000004": {"id": "recUUsUChvL3yQ96g", "name": "340 Mount Kemble"},
    "p0000005": {"id": "recqfxJfdqCXCLOuD", "name": "61 S Paramus"},
    "p0000006": {"id": "recxF4R64gbb5Sowj", "name": "575 Broadway"},
    "p0000007": {"id": "recF3zFKbY4wJ4P40", "name": "1700 East Putnam"},
    # 60-18 Metropolitan & 132-40 Metropolitan are passive/non-Yardi; add Yardi
    # codes here if/when accountant starts publishing those statements.
}

# Folder-name → property_id (used when walking the Dropbox tree directly)
FOLDER_TO_PROPERTY: dict[str, str] = {
    "paramus plaza": "recQX1kpeJKqIzvkU",
    "340 mt kemble morristown": "recUUsUChvL3yQ96g",
    "340 mount kemble": "recUUsUChvL3yQ96g",
    "61 s paramus": "recqfxJfdqCXCLOuD",
    "575 broadway": "recxF4R64gbb5Sowj",
    "1700 east putnam greenwich": "recF3zFKbY4wJ4P40",
    "1700 east putnam": "recF3zFKbY4wJ4P40",
}

# Folder names to skip entirely during discovery (no Yardi data, passive deals, etc.)
FOLDER_SKIPLIST = [
    "0. overall property management",
    "41 flatbush",
    "red bank - 1 river centre",
    "red bank",
]


# ─────────────────────────────────────────────────────────────────────────
# Account-name → GL code overrides (for aliases Yardi uses)
# ─────────────────────────────────────────────────────────────────────────
# Anything in the Yardi file that doesn't match gl-accounts.json by exact name
# (case-insensitive, punctuation/whitespace normalized) maps through this table.
NAME_OVERRIDES: dict[str, str] = {
    # ─── Income recoveries ─────────────────────────────────────────
    "property tax recovery": "4114",         # rolled into Est. Property Tax
    "cam reconciliation": "4111",
    "estimated property tax reconciliation": "4126",
    "tenant work order income": "4703",
    "bank interest": "4820",

    # ─── Contract Services (indoor/outdoor variants) ───────────────
    "elevator contract (o)": "5483",
    "elevator contract (i)": "5483",
    "esc/elevator contract (i)": "5483",
    "esc/elevator contract (o)": "5483",
    "other esc/elevator (o)": "5486",
    "elevator repairs (i)": "5486",
    "elevator repairs (o)": "5486",
    "elevator phone lines": "5484",
    "landscaping contract (o)": "5525",
    "security contract (o)": "5504",
    "security contract (i)": "5504",
    "cctv/ access control": "5504",
    "fire system contract": "5505",
    "fire system inspections & testing": "5534",
    "fire extinguisher annual inspection": "5506",
    "plot/sidewalk swp cont (o)": "5530",
    "snow removal": "5530",

    # ─── R&M (variants) ────────────────────────────────────────────
    "other cleaning related (o)": "5454",
    "other cleaning related (i)": "5454",
    "other building & related (o)": "5497",
    "other building & related (i)": "5497",
    "repair parking lot (o)": "5497",
    "roof - repairs (o)": "5509",
    "roof repairs (o)": "5509",
    "interior landscape contract": "5493",
    "meter readings": "5497",
    "tenant maint": "5421",
    "tenant reimbursement expense (wo)": "5421",

    # ─── Utilities recoveries (passthru) ───────────────────────────
    "recoverable electricity": "5603",
    "recoverable gas": "5606",
    "recoverable water & sewer": "5608",

    # ─── Insurance ─────────────────────────────────────────────────
    "insurance - property": "6523",
    "insurance - liability": "6522",
    "insurance - other": "6525",

    # ─── Real estate taxes ─────────────────────────────────────────
    "property tax": "6512",
    "real estate tax": "6512",

    # ─── Management fees ───────────────────────────────────────────
    "management fee": "6501",
    "management fees": "6501",

    # ─── G&A ───────────────────────────────────────────────────────
    "tax preparation fees": "7071",
    "financial statement fees": "7071",
    "accounting/audit": "7071",
    "administrative expenses": "7066",
    "legal leasing fees": "7068",
    "other professional fees": "7064",
    "professional fees": "7064",
    "yardi subscription": "7070",
    "bank fees": "7069",
    "banking": "7069",

    # ─── Debt / D&A ────────────────────────────────────────────────
    "mortgage interest": "7151",
    "1st mortgage interest": "7152",
    "2nd mortgage interest": "7153",
    "depreciation expense": "7210",
    "amortization expense": "7220",
}

# Intermediate Yardi sub-totals (e.g. "TOTAL CONTRACT SERVICES - OUTDOOR")
# are not part of our COA. Skip them — the actual leaf-account values roll up
# naturally when we sum by GL code in the UI.
SKIP_PATTERNS: list[re.Pattern] = [
    re.compile(r"^\s*total\s+contract\s+services", re.I),
    re.compile(r"^\s*total\s+r&m\s*-\s*(outdoor|indoor)", re.I),
    re.compile(r"^\s*total\s+direct\s+recoverable", re.I),
    re.compile(r"^\s*total\s+non-recoverable\s+utlities", re.I),  # Yardi typo
    re.compile(r"^\s*total\s+non-recoverable\s+utilities", re.I),
    re.compile(r"^\s*total\s+other\s+expenses", re.I),
    re.compile(r"^\s*net\s+cash\s+flow", re.I),
]

# Section totals we still want to carry through to actuals (labels we keep
# even if not a leaf). These Yardi labels map to GL codes that exist in our COA.
TOTAL_NAME_TO_GL: dict[str, str] = {
    "total commercial rent income": "4090",
    "total rental income/recoverable": "4190",
    "total income": "4990",
    "total other income": "4990",
    "total cleaning": "5460",
    "total repairs & maintenance": "5550",
    "total utilities": "5610",
    "total administrative expenses - recoverable": "5660",
    "total management fee": "6510",
    "total management fees": "6510",
    "total real estate taxes": "6520",
    "total insurance": "6530",
    "total building expenses": "6550",
    "total operating expenses": "6998",
    "net operating income": "6999",
    "total non-recoverable expeneses - g&a": "7099",
    "total non-recoverable expenses - g&a": "7099",
    "total general & administrative expenses": "7099",
    "total capital expenses": "7149",
    "total debt service": "7199",
    "total depreciation & amortization expense": "7299",
    "total depreciation & amortization": "7299",
    "total construction costs": "8099",
}


# ─────────────────────────────────────────────────────────────────────────
# Helpers
# ─────────────────────────────────────────────────────────────────────────
MONTHS = [
    "jan", "feb", "mar", "apr", "may", "jun",
    "jul", "aug", "sep", "oct", "nov", "dec",
]


def normalize(s: str) -> str:
    """Normalize an account name for map lookup."""
    if s is None:
        return ""
    s = str(s).strip().lower()
    # Collapse whitespace, remove redundant punctuation spacing
    s = re.sub(r"\s+", " ", s)
    # Normalize a few punctuation variants
    s = s.replace(" :", ":").replace(": ", ":").replace(":", ": ")
    s = re.sub(r"\s+", " ", s).strip()
    return s


def load_name_to_gl() -> dict[str, str]:
    """Build name→GL code map from gl-accounts.json, merged with overrides."""
    gl_path = DATA_DIR / "gl-accounts.json"
    with open(gl_path) as f:
        gl = json.load(f)
    m: dict[str, str] = {}
    for row in gl:
        name = normalize(row.get("name") or "")
        code = str(row.get("code") or "").strip()
        if name and code:
            m.setdefault(name, code)
    # Merge totals
    for k, v in TOTAL_NAME_TO_GL.items():
        m[normalize(k)] = v
    # Merge overrides (highest priority)
    for k, v in NAME_OVERRIDES.items():
        m[normalize(k)] = v
    return m


def load_gl_accounts_by_code() -> dict[str, dict[str, Any]]:
    with open(DATA_DIR / "gl-accounts.json") as f:
        gl = json.load(f)
    return {str(g["code"]): g for g in gl}


def read_supabase_config() -> tuple[str, str]:
    """Read SUPABASE_URL and SUPABASE_KEY from config.js."""
    env_url = os.environ.get("SUPABASE_URL")
    env_key = os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_SERVICE_KEY")
    if env_url and env_key:
        return env_url.rstrip("/"), env_key
    if not CONFIG_JS.exists():
        sys.exit(
            "config.js not found. Create it or set SUPABASE_URL + SUPABASE_KEY env vars."
        )
    text = CONFIG_JS.read_text()
    url_m = re.search(r"SUPABASE_URL\s*=\s*['\"]([^'\"]+)['\"]", text)
    key_m = re.search(r"SUPABASE_KEY\s*=\s*['\"]([^'\"]+)['\"]", text)
    if not (url_m and key_m):
        sys.exit("Could not parse SUPABASE_URL / SUPABASE_KEY from config.js")
    return url_m.group(1).rstrip("/"), key_m.group(1)


def find_dropbox_root() -> Path | None:
    for p in DROPBOX_ROOTS:
        if p.exists() and p.is_dir():
            return p
    return None


def find_property_folder(root: Path, property_name_hint: str) -> Path | None:
    propmgmt = root / PROPMGMT_DIRNAME
    if not propmgmt.exists():
        return None
    hint = property_name_hint.lower()
    for child in propmgmt.iterdir():
        if not child.is_dir():
            continue
        if hint in child.name.lower() or child.name.lower() in hint:
            return child
    return None


def find_accounting_folder(prop_folder: Path) -> Path | None:
    accounting = prop_folder / ACCOUNTING_DIRNAME
    if not accounting.exists():
        # try case-insensitive match
        for c in prop_folder.iterdir():
            if c.is_dir() and c.name.lower() == ACCOUNTING_DIRNAME.lower():
                accounting = c
                break
        else:
            return None
    # Look inside for the month-financials subfolder
    candidates = [c for c in accounting.iterdir() if c.is_dir()]
    for name in MONTH_FINANCIALS_DIRNAMES:
        for c in candidates:
            if c.name.lower() == name.lower():
                return c
    # Fallback: first folder starting with "A " or "A -"
    for c in candidates:
        if c.name.lower().startswith(("a ", "a-", "a.")):
            return c
    return None


def all_income_statements(folder: Path) -> list[Path]:
    """Return every income-statement xlsx in `folder` (recursive)."""
    patterns = [
        "12Monthsincomestatement*.xlsx",
        "12MonthsIncomeStatement*.xlsx",
        "Income_Statement*.xlsx",
        "income_statement*.xlsx",
        "*incomestatement*.xlsx",
        "*IncomeStatement*.xlsx",
        "*Income Statement*.xlsx",
        "*income statement*.xlsx",
        "*Income_Statement*.xlsx",
        "*income_statement*.xlsx",
    ]
    seen: set[str] = set()
    out: list[Path] = []
    for pat in patterns:
        for p in glob.glob(str(folder / pat)):
            if p not in seen:
                seen.add(p)
                out.append(Path(p))
        for p in glob.glob(str(folder / "**" / pat), recursive=True):
            if p not in seen:
                seen.add(p)
                out.append(Path(p))
    # Skip Excel lock files (created when a file is open in Excel)
    out = [p for p in out if not p.name.startswith("~$")]
    return out


def latest_income_statement(folder: Path, debug: bool = False) -> Path | None:
    """Return the most recent 12-month income statement xlsx in `folder`.

    Searches recursively (accountants sometimes nest files in year/month
    subfolders). Matches any xlsx whose name contains 'incomestatement'
    (case-insensitive) plus common Yardi prefix variants.
    """
    patterns = [
        # 12-month Yardi export (what was in data/2025 Actuals/)
        "12Monthsincomestatement*.xlsx",
        "12MonthsIncomeStatement*.xlsx",
        # Accountant's monthly folders (2026+): "Income_Statement 01.2026.xlsx"
        "Income_Statement*.xlsx",
        "income_statement*.xlsx",
        # Generic catch-all
        "*incomestatement*.xlsx",
        "*IncomeStatement*.xlsx",
        "*Income Statement*.xlsx",
        "*income statement*.xlsx",
        "*Income_Statement*.xlsx",
        "*income_statement*.xlsx",
    ]
    seen: set[str] = set()
    matches: list[str] = []
    for pat in patterns:
        # Non-recursive first (preferred — top of folder)
        for p in glob.glob(str(folder / pat)):
            if p not in seen:
                seen.add(p)
                matches.append(p)
        # Recursive fallback — any subfolder
        for p in glob.glob(str(folder / "**" / pat), recursive=True):
            if p not in seen:
                seen.add(p)
                matches.append(p)

    if not matches:
        if debug:
            contents = [c.name for c in folder.iterdir()][:15]
            print(f"      (folder contents: {contents})")
            # Report ALL xlsx files found recursively to help diagnose naming
            all_xlsx = sorted(glob.glob(str(folder / "**" / "*.xlsx"), recursive=True))
            if all_xlsx:
                print(f"      (all xlsx recursively, up to 10):")
                for p in all_xlsx[:10]:
                    rel = os.path.relpath(p, folder)
                    print(f"        {rel}")
                if len(all_xlsx) > 10:
                    print(f"        … and {len(all_xlsx) - 10} more")
            else:
                print("      (no xlsx files found at any depth)")
        return None

    matches.sort(key=lambda p: os.path.getmtime(p), reverse=True)
    return Path(matches[0])


def materialize_dropbox_file(path: Path, timeout: int = 300) -> bool:
    """
    Ensure a Dropbox Online-Only file is fully downloaded locally.
    On macOS Dropbox uses the File Provider extension — a full file read
    blocks until the File Provider materializes the file. We use `cat > /dev/null`
    which is the most reliable trigger.

    Returns True if the file is ready, False on timeout / failure.
    """
    try:
        if path.stat().st_size > 0:
            return True
    except OSError:
        return False

    import subprocess

    # Running `cat` blocks until File Provider supplies the full contents.
    # This is more reliable than partial reads because some FP extensions
    # won't start materialization for small reads.
    try:
        result = subprocess.run(
            ["/bin/cat", str(path)],
            stdout=subprocess.DEVNULL,
            stderr=subprocess.DEVNULL,
            timeout=timeout,
            check=False,
        )
    except subprocess.TimeoutExpired:
        return path.stat().st_size > 0 if path.exists() else False
    except Exception:
        pass

    # Final check
    try:
        return path.stat().st_size > 0
    except OSError:
        return False


# ─────────────────────────────────────────────────────────────────────────
# xlsx parsers
# ─────────────────────────────────────────────────────────────────────────
MONTH_ABBR = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}


def parse_period_from_name(filename: str) -> tuple[int, int] | None:
    """Best-effort extraction of (year, period_end_month) from a filename.

    Handles variants we've observed:
      "Income_Statement 03.31.26.xlsx"     → (2026, 3)
      "Income_Statement 03.31.2026.xlsx"   → (2026, 3)
      "Income_Statement 01.2026.xlsx"      → (2026, 1)
      "Income_Statement Mar 2026.xlsx"     → (2026, 3)
      "Income_Statement Q2 2025.xlsx"      → (2025, 6)
      "Income_Statement Dec 2025.xlsx"     → (2025, 12)
      "Income_Statement 2026-03-31.xlsx"   → (2026, 3)
    """
    s = filename
    # Q1/Q2/Q3/Q4 YYYY
    m = re.search(r"\bQ([1-4])\s+(\d{4})", s, re.I)
    if m:
        q = int(m.group(1))
        return (int(m.group(2)), q * 3)
    # Month-name YYYY
    m = re.search(r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\w*\s+(\d{4})", s, re.I)
    if m:
        return (int(m.group(2)), MONTH_ABBR[m.group(1).lower()[:3]])
    # MM.DD.YY or MM.DD.YYYY
    m = re.search(r"\b(\d{1,2})[\.\-/](\d{1,2})[\.\-/](\d{2,4})\b", s)
    if m:
        mo = int(m.group(1))
        yr = int(m.group(3))
        yr = yr if yr >= 1900 else 2000 + yr
        if 1 <= mo <= 12 and yr >= 2020:
            return (yr, mo)
    # MM.YYYY or M.YYYY
    m = re.search(r"\b(\d{1,2})\.(\d{4})\b", s)
    if m:
        mo = int(m.group(1))
        yr = int(m.group(2))
        if 1 <= mo <= 12:
            return (yr, mo)
    # YYYY-MM-DD
    m = re.search(r"(\d{4})-(\d{1,2})-(\d{1,2})", s)
    if m:
        return (int(m.group(1)), int(m.group(2)))
    return None


def db_latest_month(url: str, key: str, property_id: str, year: int) -> int:
    """Return max month already loaded in actuals_line_items for (property, year)."""
    endpoint = f"{url}/rest/v1/actuals_line_items"
    headers = {"apikey": key, "Authorization": f"Bearer {key}"}
    params = {
        "property_id": f"eq.{property_id}",
        "year": f"eq.{year}",
        "select": "month",
        "order": "month.desc",
        "limit": "1",
    }
    try:
        resp = requests.get(endpoint, headers=headers, params=params, timeout=30)
        if resp.status_code == 200:
            rows = resp.json()
            if rows:
                return int(rows[0]["month"])
    except Exception:
        pass
    return 0


def _detect_format(all_rows: list[tuple]) -> str:
    """Return 'twelve_month', 'ptd_ytd', or 'unknown'."""
    for row in all_rows[:15]:
        if not row:
            continue
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " | ".join(cells)
        if "period to date" in joined and "year to date" in joined:
            return "ptd_ytd"
        if cells and cells[0] == "account":
            # 12-month format has "account" column header followed by month columns
            return "twelve_month"
    return "unknown"


def parse_ptd_ytd(path: Path) -> dict[str, Any]:
    """Parse Yardi 'Income Statement' export (GL codes in col A, PTD/YTD in col C/E).

    Returns dict including:
        property_code, property_name, year,
        period_start_month, period_end_month,
        gl_ytd: { gl_code: ytd_amount },
        gl_ptd: { gl_code: ptd_amount },
        gl_account: { gl_code: account_name },
        source_file: str.
    """
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    property_code = property_name = None
    year = None
    period_start_month = period_end_month = None
    header_row_idx = None

    for i, row in enumerate(all_rows):
        if not row:
            continue
        # Check for the PTD/YTD header row first — this one may have col 0 empty
        cells = [str(c).strip().lower() if c is not None else "" for c in row]
        joined = " | ".join(cells)
        if "period to date" in joined and "year to date" in joined:
            header_row_idx = i
            break
        if row[0] is None:
            continue
        c0 = str(row[0]).strip()
        # Row 1: property + code, or just the filename — try to parse
        m_prop = re.search(r"(.+?)\s*\(([pP]\d+)\)", c0)
        if m_prop and not property_code:
            property_name = m_prop.group(1).strip()
            property_code = m_prop.group(2).strip().lower()
            continue
        # Period row
        if c0.lower().startswith("period"):
            # "Period = Jan 2026-Mar 2026" or "Period = Feb 2026" etc.
            m = re.search(
                r"(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4})"
                r"\s*-?\s*"
                r"(?:(Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)\s*(\d{4}))?",
                c0,
                re.I,
            )
            if m:
                period_start_month = MONTH_ABBR[m.group(1).lower()]
                year = int(m.group(2))
                if m.group(3):
                    period_end_month = MONTH_ABBR[m.group(3).lower()]
                else:
                    period_end_month = period_start_month
            continue
    if header_row_idx is None:
        raise ValueError(f"Could not locate PTD/YTD header row in {path.name}")
    if not year:
        raise ValueError(f"Could not parse year from {path.name}")
    if not property_code:
        # Filename sometimes contains the code, e.g. "...p0000005..."
        m = re.search(r"(p\d{7})", path.name, re.I)
        if m:
            property_code = m.group(1).lower()
    if not property_code:
        raise ValueError(f"Could not parse property code from {path.name}")

    # Detect which column holds PTD $ and YTD $. Usually col 2 and 4.
    header = all_rows[header_row_idx]
    ptd_col = ytd_col = None
    for j, c in enumerate(header):
        if c is None:
            continue
        s = str(c).strip().lower()
        if ptd_col is None and "period to date" in s:
            ptd_col = j
        elif ytd_col is None and "year to date" in s:
            ytd_col = j
    if ptd_col is None or ytd_col is None:
        raise ValueError(f"Could not locate PTD/YTD columns in {path.name}")

    gl_ytd: dict[str, float] = {}
    gl_ptd: dict[str, float] = {}
    gl_account: dict[str, str] = {}

    for row in all_rows[header_row_idx + 1 :]:
        if not row or row[0] is None:
            continue
        gl_raw = str(row[0]).strip()
        # Expect a 4-digit GL code in col 0
        m = re.match(r"^\d{4}$", gl_raw)
        if not m:
            continue
        gl = gl_raw
        # Skip Head lines (no amount) — those have no numeric values
        acct_name = (str(row[1]).strip() if len(row) > 1 and row[1] is not None else "")
        # Strip leading indentation
        acct_name = re.sub(r"^\s+", "", acct_name)
        try:
            ptd = float(row[ptd_col]) if row[ptd_col] not in (None, "") else None
        except (TypeError, ValueError):
            ptd = None
        try:
            ytd = float(row[ytd_col]) if row[ytd_col] not in (None, "") else None
        except (TypeError, ValueError):
            ytd = None
        if ptd is None and ytd is None:
            continue  # header-only line
        if ptd is not None:
            gl_ptd[gl] = ptd
        if ytd is not None:
            gl_ytd[gl] = ytd
        gl_account[gl] = acct_name

    return {
        "property_code": property_code,
        "property_name": property_name,
        "year": year,
        "period_start_month": period_start_month,
        "period_end_month": period_end_month,
        "gl_ytd": gl_ytd,
        "gl_ptd": gl_ptd,
        "gl_account": gl_account,
        "source_file": str(path),
    }


def parse_income_statement(
    path: Path,
    name_to_gl: dict[str, str],
) -> dict[str, Any]:
    """
    Parse a Yardi income statement (12-month format OR PTD/YTD format).

    12-month format returns:
        property_code, property_name, year,
        rows: [{ account_name, gl_code, months: {1..12: float} }],
        unmatched: [...],
        _format: 'twelve_month'

    PTD/YTD format returns:
        property_code, property_name, year,
        period_start_month, period_end_month,
        gl_ytd, gl_ptd, gl_account,
        _format: 'ptd_ytd'

    The caller should inspect result['_format'] to branch on shape.
    """
    # Quick peek at first ~15 rows to detect format
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    peek = list(ws.iter_rows(values_only=True, max_row=15))
    fmt = _detect_format(peek)
    wb.close()

    if fmt == "ptd_ytd":
        out = parse_ptd_ytd(path)
        out["_format"] = "ptd_ytd"
        return out

    # Fall through to 12-month parser (legacy Yardi export)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    all_rows = list(ws.iter_rows(values_only=True))

    # Header parsing
    property_code = None
    property_name = None
    year = None
    month_headers: list[int] = []  # zero-based column index -> month number (1..12)

    header_row_idx = None
    for i, row in enumerate(all_rows):
        if not row or row[0] is None:
            continue
        cell0 = str(row[0]).strip()
        if cell0.lower().startswith("property"):
            m = re.search(r"Property\s*=\s*(.+?)\s*\(([^)]+)\)", cell0)
            if m:
                property_name = m.group(1).strip()
                property_code = m.group(2).strip()
        elif cell0.lower().startswith("period"):
            m = re.search(r"(\d{1,2})/(\d{4})", cell0)
            if m:
                year = int(m.group(2))
        elif cell0.strip().upper() == "ACCOUNT":
            header_row_idx = i
            # Find month columns
            for j, c in enumerate(row):
                if c is None:
                    continue
                s = str(c).strip().lower()
                for mi, mname in enumerate(MONTHS, start=1):
                    if s.startswith(mname):
                        month_headers.append((j, mi))
                        break
            break

    if header_row_idx is None or not month_headers:
        raise ValueError(f"Could not locate header row / month columns in {path.name}")
    if not property_code:
        raise ValueError(f"Could not parse property code from {path.name}")
    if not year:
        raise ValueError(f"Could not parse year from {path.name}")

    # Data rows: labels are indented. We'll keep LEAF rows (those with values)
    # AND section-total rows whose names resolve to a GL code via TOTAL_NAME_TO_GL.
    parsed_rows: list[dict[str, Any]] = []
    unmatched: list[str] = []

    for row in all_rows[header_row_idx + 1 :]:
        if not row or row[0] is None:
            continue
        label_raw = str(row[0])
        label = label_raw.strip()
        if not label:
            continue

        months_dict: dict[int, float] = {}
        has_any_value = False
        for col_idx, month_num in month_headers:
            val = row[col_idx] if col_idx < len(row) else None
            if val is None or val == "":
                continue
            try:
                fv = float(val)
            except (TypeError, ValueError):
                continue
            if fv == 0:
                # Keep zero if there's any non-zero month for this row; we'll
                # post-filter below.
                months_dict[month_num] = 0.0
                continue
            months_dict[month_num] = fv
            has_any_value = True

        if not has_any_value:
            # No data on this row (section header). Skip.
            continue

        # Skip intermediate Yardi sub-totals that don't map to our COA
        if any(pat.match(label) for pat in SKIP_PATTERNS):
            continue

        norm = normalize(label)
        gl_code = name_to_gl.get(norm)

        if not gl_code:
            # Try fuzzy variants
            alt = norm.replace("expeneses", "expenses")
            gl_code = name_to_gl.get(alt)
        if not gl_code:
            unmatched.append(label)
            continue

        parsed_rows.append(
            {
                "account_name": label,
                "gl_code": gl_code,
                "months": months_dict,
            }
        )

    return {
        "property_code": property_code,
        "property_name": property_name,
        "year": year,
        "rows": parsed_rows,
        "unmatched": unmatched,
        "source_file": str(path),
        "_format": "twelve_month",
    }


def derive_monthly_from_ptd_files(parsed_files: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Given a chronologically-sorted list of parsed PTD/YTD files (all for the
    same property+year), derive per-month values per GL code.

    Strategy:
      For each GL code, walk the files in order. The value for months
      (period_start..period_end) of each file is this file's PTD (period-to-date)
      divided evenly across those months. This handles:
        • single-month files → PTD goes straight into that month
        • quarterly files (e.g. Jan-Mar) → PTD split evenly across Jan,Feb,Mar
        • mixed (e.g. Jan one file + Feb-Mar another) → precise Jan plus
          evenly-split Feb,Mar
      Later files for overlapping months OVERWRITE earlier ones (assumption:
      newer = more accurate close).

    Returns:
        {
          'property_code', 'property_name', 'year',
          'rows': [ {gl_code, account_name, months: {1..N: float}} ],
          'months_covered': set of month numbers that have data,
        }
    """
    if not parsed_files:
        return {"rows": [], "months_covered": set()}

    property_code = parsed_files[-1]["property_code"]
    property_name = parsed_files[-1]["property_name"]
    year = parsed_files[-1]["year"]

    # gl_code → month_num → amount
    gl_months: dict[str, dict[int, float]] = {}
    gl_account: dict[str, str] = {}
    months_covered: set[int] = set()

    # Sort so that the most-precise (shortest) periods are processed first.
    # Primary: period_end ascending. Secondary: period_start descending (so a
    # single-month Mar file is handled before a full Q1 file that also ends
    # in March).
    files_sorted = sorted(
        parsed_files,
        key=lambda p: (
            p.get("period_end_month") or 0,
            -(p.get("period_start_month") or 0),
        ),
    )

    for pf in files_sorted:
        ps = pf.get("period_start_month")
        pe = pf.get("period_end_month")
        if not ps or not pe:
            continue
        period_months = list(range(ps, pe + 1))
        for gl, ptd in pf["gl_ptd"].items():
            slot = gl_months.setdefault(gl, {})
            unfilled = [m for m in period_months if m not in slot]
            if not unfilled:
                continue  # all months already have precise values
            # Subtract months we already know to get the remainder for the unfilled ones
            already_total = sum(slot.get(m, 0.0) for m in period_months if m in slot)
            remaining = ptd - already_total
            share = remaining / len(unfilled)
            for m in unfilled:
                slot[m] = share
                months_covered.add(m)
        for gl, acct in pf["gl_account"].items():
            if gl not in gl_account or (acct and len(acct) > len(gl_account[gl])):
                gl_account[gl] = acct

    rows = []
    for gl, months in gl_months.items():
        rows.append({
            "gl_code": gl,
            "account_name": gl_account.get(gl, ""),
            "months": months,
        })

    return {
        "property_code": property_code,
        "property_name": property_name,
        "year": year,
        "rows": rows,
        "months_covered": months_covered,
    }


# ─────────────────────────────────────────────────────────────────────────
# Supabase upsert
# ─────────────────────────────────────────────────────────────────────────
def supabase_upsert(
    url: str,
    key: str,
    property_id: str,
    year: int,
    rows: list[dict[str, Any]],
    gl_by_code: dict[str, dict[str, Any]],
    verbose: bool = True,
) -> int:
    """
    Upsert rows into actuals_line_items. Conflict key:
        (property_id, year, gl_code, month)

    Strategy: first DELETE all existing rows for this (property_id, year), then INSERT.
    That gives us clean "latest snapshot wins" semantics and matches the fact
    that the accountant re-publishes the full 12-month file each month.
    """
    endpoint = f"{url}/rest/v1/actuals_line_items"
    headers = {
        "apikey": key,
        "Authorization": f"Bearer {key}",
        "Content-Type": "application/json",
        "Prefer": "return=minimal",
    }

    # 1) Delete existing rows for this property/year
    del_resp = requests.delete(
        endpoint,
        headers=headers,
        params={"property_id": f"eq.{property_id}", "year": f"eq.{year}"},
        timeout=60,
    )
    if del_resp.status_code not in (200, 204):
        raise RuntimeError(
            f"DELETE failed ({del_resp.status_code}): {del_resp.text[:400]}"
        )
    if verbose:
        print(f"    cleared existing actuals for {property_id} / {year}")

    # 2) Build payload
    payload: list[dict[str, Any]] = []
    for r in rows:
        gl = r["gl_code"]
        acct = gl_by_code.get(gl, {}).get("name") or r["account_name"]
        for month, amount in sorted(r["months"].items()):
            payload.append(
                {
                    "property_id": property_id,
                    "year": year,
                    "gl_code": gl,
                    "account_name": acct,
                    "month": month,
                    "amount": round(float(amount), 2),
                }
            )
    if not payload:
        return 0

    # 3) Insert in chunks
    chunk = 500
    inserted = 0
    for i in range(0, len(payload), chunk):
        part = payload[i : i + chunk]
        resp = requests.post(endpoint, headers=headers, json=part, timeout=120)
        if resp.status_code not in (200, 201, 204):
            raise RuntimeError(
                f"POST failed ({resp.status_code}): {resp.text[:400]}"
            )
        inserted += len(part)
    return inserted


# ─────────────────────────────────────────────────────────────────────────
# File discovery
# ─────────────────────────────────────────────────────────────────────────
def discover_files_grouped(
    explicit_file: str | None,
    only_property: str | None,
) -> dict[str, list[Path]]:
    """
    Return a mapping of property-folder-name → [xlsx paths].
    Each list contains every monthly/quarterly income-statement file for that
    property, unordered (caller sorts by parsed period end).

    If --file is given, returns a single-entry map keyed by the filename.
    """
    if explicit_file:
        p = Path(explicit_file).expanduser().resolve()
        if not p.exists():
            sys.exit(f"File not found: {p}")
        return {p.name: [p]}

    grouped: dict[str, list[Path]] = {}
    root = find_dropbox_root()
    if root is None:
        print(
            "⚠ Dropbox root not found. Falling back to local data/ folder.",
            file=sys.stderr,
        )
        fallback = sorted((DATA_DIR / "2025 Actuals").glob("*.xlsx"))
        if fallback:
            grouped["(local fallback)"] = list(fallback)
        return grouped

    propmgmt = root / PROPMGMT_DIRNAME
    if not propmgmt.exists():
        print(f"⚠ '{PROPMGMT_DIRNAME}' not found directly. Searching in {root} …")
        for child in root.iterdir():
            if child.is_dir() and "property management" in child.name.lower():
                print(f"  found candidate: {child.name}")
                propmgmt = child
                break
        else:
            print(f"Contents of {root}:")
            for c in sorted(root.iterdir()):
                print(f"  {c.name}")
            sys.exit(f"\nCould not locate a property management folder in {root}.")

    print(f"✓ Using property-management root: {propmgmt}")
    for child in sorted(propmgmt.iterdir()):
        if not child.is_dir():
            continue
        if any(skip in child.name.lower() for skip in FOLDER_SKIPLIST):
            print(f"  ⏭  {child.name}: skiplisted")
            continue
        acct = find_accounting_folder(child)
        if not acct:
            print(f"  ⏭  {child.name}: no 'A - Month Quarter Financials' folder found")
            continue
        stmts = all_income_statements(acct)
        if not stmts:
            try:
                contents = [c.name for c in acct.iterdir()][:15]
            except OSError:
                contents = []
            print(f"  ⏭  {child.name}: no income statements in '{acct.name}'")
            if contents:
                print(f"      (folder contents: {contents})")
            continue
        if only_property:
            stmts = [s for s in stmts if only_property.lower() in s.name.lower()
                     or only_property.lower() in child.name.lower()]
            if not stmts:
                continue
        # Pick the file with the NEWEST period (not newest mtime). Mtime
        # can be misleading if you re-download an older file. We prefer
        # filenames with a parseable (year, month), newest period winning.
        def _sort_key(p):
            period = parse_period_from_name(p.name)
            if period:
                return (1, period[0], period[1], os.path.getmtime(p))
            return (0, 0, 0, os.path.getmtime(p))
        stmts.sort(key=_sort_key, reverse=True)
        latest = stmts[0]
        print(f"  ✓ {child.name}: using latest → {latest.name}  ({len(stmts)} total found)")
        grouped[child.name] = [latest]

    return grouped


# ─────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────
def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument(
        "--commit",
        action="store_true",
        help="Actually write to Supabase. Default is dry-run.",
    )
    ap.add_argument("--year", type=int, help="Force year (default: parsed from file).")
    ap.add_argument(
        "--min-year",
        type=int,
        default=2026,
        help="Skip files older than this year. 2026 by default (pre-2026 files use old COA).",
    )
    ap.add_argument("--property", help="Yardi property code or name substring.")
    ap.add_argument("--file", help="Process a single xlsx file and stop.")
    ap.add_argument("--verbose", "-v", action="store_true")
    ap.add_argument(
        "--force-refresh",
        action="store_true",
        help="Reload files even if their period is already in Supabase.",
    )
    args = ap.parse_args()

    name_to_gl = load_name_to_gl()
    gl_by_code = load_gl_accounts_by_code()

    grouped = discover_files_grouped(args.file, args.property)
    if not grouped:
        print("No files found. Check Dropbox path or use --file.")
        return 1

    total_files = sum(len(v) for v in grouped.values())
    print(f"\nFound {total_files} file(s) across {len(grouped)} propert(y/ies).")

    supabase_url = supabase_key = None
    if args.commit:
        supabase_url, supabase_key = read_supabase_config()

    any_unmatched = False
    total_inserted = 0
    total_files_processed = 0

    for folder_name, file_paths in grouped.items():
        print(f"\n━━━ {folder_name} ({len(file_paths)} file(s)) ━━━")

        # ── Freshness check: skip files whose period is already loaded ──
        # Map folder name → property_id so we can query what's in DB
        property_id_for_folder = None
        for key, pid in FOLDER_TO_PROPERTY.items():
            if key in folder_name.lower() or folder_name.lower() in key:
                property_id_for_folder = pid
                break

        if property_id_for_folder and not args.file:
            # Peek at each file's filename to get a (year, month) hint
            fresh: list[Path] = []
            for f in file_paths:
                period = parse_period_from_name(f.name)
                if not period:
                    # Can't tell from filename — let it through, parser will decide
                    fresh.append(f)
                    continue
                yr, mo = period
                if yr < args.min_year:
                    print(f"  ⏭  {f.name}: year {yr} < --min-year {args.min_year}")
                    continue
                if args.commit or args.force_refresh:
                    db_max = 0
                    if args.commit:
                        db_max = db_latest_month(supabase_url, supabase_key, property_id_for_folder, yr)
                    if db_max >= mo and not args.force_refresh:
                        print(f"  ⏭  {f.name}: period {yr}-{mo:02d} already loaded (db has thru month {db_max})")
                        continue
                fresh.append(f)
            if not fresh:
                print("  ✓ up to date — nothing new to load")
                continue
            file_paths = fresh

        # Materialize every file first. Drop any that fail to download.
        usable: list[Path] = []
        for f in file_paths:
            try:
                sz = f.stat().st_size
            except OSError:
                sz = 0
            if sz == 0:
                print(f"  … pulling down {f.name} from Dropbox …")
                if not materialize_dropbox_file(f):
                    print(f"  ✗ timed out downloading {f.name} — skipping")
                    continue
                print(f"  ✓ materialized ({f.stat().st_size:,} bytes)")
            usable.append(f)
        file_paths = usable
        if not file_paths:
            print("  ⏭  no files available to parse")
            continue

        # Parse all files for this property
        parsed_by_year: dict[int, dict[str, list[dict[str, Any]]]] = {}
        # structure: year → { 'ptd_ytd': [parsed, ...], 'twelve_month': [parsed, ...] }
        for f in file_paths:
            try:
                parsed = parse_income_statement(f, name_to_gl)
            except Exception as e:
                print(f"  ✗ {f.name}: parse error: {e}")
                continue
            yr = args.year or parsed.get("year")
            if not yr:
                print(f"  ✗ {f.name}: no year parsed, skipping")
                continue
            if yr < args.min_year:
                continue
            fmt = parsed.get("_format", "unknown")
            parsed_by_year.setdefault(yr, {}).setdefault(fmt, []).append(parsed)
            total_files_processed += 1

        if not parsed_by_year:
            print(f"  ⏭  no files in scope (min-year={args.min_year})")
            continue

        # Process each year independently
        for yr, by_fmt in sorted(parsed_by_year.items()):
            ptd_list = by_fmt.get("ptd_ytd", [])
            twelve_list = by_fmt.get("twelve_month", [])

            # Prefer 12-month if available (it's already per-month)
            if twelve_list:
                parsed = max(twelve_list, key=lambda p: os.path.getmtime(p["source_file"]))
                prop_code = parsed["property_code"]
                prop_name = parsed["property_name"]
                rows = parsed["rows"]
                unmatched = parsed.get("unmatched", [])
                months_covered = sorted({m for r in rows for m in r["months"].keys()})
            elif ptd_list:
                # Sort by period_end then mtime
                ptd_list.sort(key=lambda p: (
                    p.get("period_end_month") or 0,
                    os.path.getmtime(p["source_file"]),
                ))
                derived = derive_monthly_from_ptd_files(ptd_list)
                prop_code = derived["property_code"]
                prop_name = derived["property_name"]
                rows = derived["rows"]
                unmatched = []
                months_covered = sorted(derived["months_covered"])
            else:
                continue

            map_entry = PROPERTY_MAP.get(prop_code)
            if not map_entry:
                print(
                    f"  ✗ Year {yr}: unknown Yardi property code {prop_code!r} "
                    f"({prop_name}). Add to PROPERTY_MAP."
                )
                continue
            supa_id = map_entry["id"]

            print(f"  ─ year {yr} ({'12-mo' if twelve_list else f'{len(ptd_list)} PTD file(s)'}) ─")
            print(f"    property  : {prop_name} ({prop_code}) → {supa_id}")
            print(f"    rows      : {len(rows)} GL codes")
            print(f"    months    : {months_covered if months_covered else '(none)'}")
            if unmatched:
                any_unmatched = True
                print(f"    ⚠ unmatched accounts ({len(unmatched)}):")
                for u in unmatched[:10]:
                    print(f"        · {u}")

            if args.verbose:
                for r in rows[:30]:
                    total = sum(r["months"].values())
                    print(f"      {r['gl_code']:>6}  {r['account_name'][:40]:40s}  total={total:>12,.2f}")

            if args.commit:
                n = supabase_upsert(
                    supabase_url, supabase_key, supa_id, yr, rows, gl_by_code
                )
                total_inserted += n
                print(f"    ✓ upserted {n} month-rows to Supabase")
            else:
                payload_count = sum(len(r["months"]) for r in rows)
                total_inserted += payload_count
                print(f"    (dry-run) would upsert {payload_count} month-rows")

    print("\n━━━ summary ━━━")
    print(f"  files parsed    : {total_files_processed}")
    print(f"  total rows      : {total_inserted}")
    print(f"  mode            : {'COMMIT' if args.commit else 'dry-run'}")
    if any_unmatched:
        print(
            "  ⚠ some account names did not match a GL code. Add them to "
            "NAME_OVERRIDES in sync_actuals.py."
        )
    if not args.commit:
        print("\nRe-run with --commit to actually write to Supabase.")
    return 0


if __name__ == "__main__":
    sys.exit(main())
