#!/usr/bin/env python3
"""
scan_prospective_deals.py — Walk 1.1 Projects - Prospective deal folders
and dump their contents to data/prospective_scan.json so Claude (running in
a separate sandbox) can read the summaries and underwrite the deals.

Usage:
    python3 scripts/scan_prospective_deals.py "9 Campus" "955 Mass Ave" "4300 Roosevelt"

    # Or scan everything in 1.1 Projects - Prospective:
    python3 scripts/scan_prospective_deals.py --all

Output file: data/prospective_scan.json

For each folder:
  - List every file with its size
  - For .xlsx: dump every sheet (first 100 rows × 20 cols) as 2D arrays
  - For .csv/.txt: dump first 5000 chars
  - For .pdf: dump first page text if pdftotext (poppler) is available;
              otherwise note "pdf_skipped"
  - For .docx: dump plain text if python-docx is installed;
              otherwise note "docx_skipped"
"""

from __future__ import annotations

import argparse
import json
import os
import subprocess
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
except ImportError:
    sys.exit("Missing dependency: pip3 install openpyxl")


SCRIPT_DIR = Path(__file__).resolve().parent
PROJECT_DIR = SCRIPT_DIR.parent
DATA_DIR = PROJECT_DIR / "data"

DROPBOX_ROOTS = [
    Path.home() / "First Mile Prop Dropbox",
    Path.home() / "Library" / "CloudStorage" / "First Mile Prop Dropbox",
]
PROSPECTIVE_DIRNAME = "1.1 Projects - Prospective"

MAX_XLSX_ROWS = 400
MAX_XLSX_COLS = 25
MAX_PDF_CHARS = 15000
MAX_DOCX_CHARS = 15000
MAX_TEXT_CHARS = 5000

SKIP_PATTERNS = ("~$", ".DS_Store", "._")


def find_prospective_root() -> Path | None:
    for p in DROPBOX_ROOTS:
        cand = p / PROSPECTIVE_DIRNAME
        if cand.exists() and cand.is_dir():
            return cand
    return None


def materialize(path: Path, timeout: int = 120) -> bool:
    try:
        if path.stat().st_size > 0:
            return True
    except OSError:
        return False
    try:
        subprocess.run(["/bin/cat", str(path)], stdout=subprocess.DEVNULL,
                       stderr=subprocess.DEVNULL, timeout=timeout, check=False)
    except Exception:
        pass
    try:
        return path.stat().st_size > 0
    except OSError:
        return False


def dump_xlsx(path: Path) -> dict[str, Any]:
    try:
        wb = openpyxl.load_workbook(path, data_only=True, read_only=True)
    except Exception as e:
        return {"error": f"openpyxl failed: {e}"}
    sheets = {}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_out = []
        for i, row in enumerate(ws.iter_rows(values_only=True)):
            if i >= MAX_XLSX_ROWS:
                break
            trimmed = list(row[:MAX_XLSX_COLS])
            # Convert datetime to str
            trimmed = [
                v.isoformat() if hasattr(v, "isoformat") else v
                for v in trimmed
            ]
            rows_out.append(trimmed)
        sheets[name] = rows_out
    wb.close()
    return {"sheets": sheets}


def dump_pdf(path: Path) -> dict[str, Any]:
    # Try pdftotext (poppler) first — usually available on Mac via homebrew
    try:
        result = subprocess.run(
            ["/usr/local/bin/pdftotext", "-layout", str(path), "-"],
            capture_output=True, timeout=60, check=False,
        )
        if result.returncode == 0 and result.stdout:
            return {"text": result.stdout.decode("utf-8", errors="replace")[:MAX_PDF_CHARS]}
    except FileNotFoundError:
        pass
    except Exception as e:
        return {"error": f"pdftotext failed: {e}"}
    # Fallback: /opt/homebrew path for Apple Silicon
    try:
        result = subprocess.run(
            ["/opt/homebrew/bin/pdftotext", "-layout", str(path), "-"],
            capture_output=True, timeout=60, check=False,
        )
        if result.returncode == 0 and result.stdout:
            return {"text": result.stdout.decode("utf-8", errors="replace")[:MAX_PDF_CHARS]}
    except FileNotFoundError:
        pass
    except Exception as e:
        return {"error": f"pdftotext (arm) failed: {e}"}
    # Fallback 2: try pypdf if available
    try:
        from pypdf import PdfReader  # type: ignore
        reader = PdfReader(str(path))
        text = "\n".join((p.extract_text() or "") for p in reader.pages[:10])
        if text.strip():
            return {"text": text[:MAX_PDF_CHARS]}
    except ImportError:
        pass
    except Exception as e:
        return {"error": f"pypdf failed: {e}"}
    return {"skipped": "no pdf backend (install poppler via 'brew install poppler' or pip3 install pypdf)"}


def dump_docx(path: Path) -> dict[str, Any]:
    try:
        from docx import Document  # type: ignore
        doc = Document(str(path))
        text = "\n".join(p.text for p in doc.paragraphs if p.text.strip())
        return {"text": text[:MAX_DOCX_CHARS]}
    except ImportError:
        return {"skipped": "python-docx not installed (pip3 install python-docx)"}
    except Exception as e:
        return {"error": f"docx read failed: {e}"}


def scan_file(path: Path) -> dict[str, Any]:
    out: dict[str, Any] = {
        "path": str(path),
        "name": path.name,
        "size": path.stat().st_size if path.exists() else 0,
    }
    name = path.name.lower()
    if any(name.startswith(p) for p in SKIP_PATTERNS):
        return None  # type: ignore
    if not materialize(path):
        out["error"] = "could not materialize from Dropbox"
        return out
    try:
        if name.endswith((".xlsx", ".xlsm")):
            out.update(dump_xlsx(path))
        elif name.endswith(".pdf"):
            out.update(dump_pdf(path))
        elif name.endswith(".docx"):
            out.update(dump_docx(path))
        elif name.endswith((".csv", ".txt", ".md")):
            out["text"] = path.read_text(errors="replace")[:MAX_TEXT_CHARS]
        else:
            out["skipped"] = f"unsupported extension"
    except Exception as e:
        out["error"] = f"{type(e).__name__}: {e}"
    return out


def scan_folder(folder: Path) -> dict[str, Any]:
    print(f"▶ scanning {folder.name} …")
    files_data = []
    for f in sorted(folder.rglob("*")):
        if not f.is_file():
            continue
        if any(f.name.startswith(p) for p in SKIP_PATTERNS):
            continue
        rel = f.relative_to(folder)
        print(f"    · {rel}")
        info = scan_file(f)
        if info:
            info["relpath"] = str(rel)
            files_data.append(info)
    return {
        "folder": str(folder),
        "name": folder.name,
        "file_count": len(files_data),
        "files": files_data,
    }


def main() -> int:
    ap = argparse.ArgumentParser(description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("deals", nargs="*", help="Deal folder names (matches by substring).")
    ap.add_argument("--all", action="store_true", help="Scan every subfolder of 1.1 Projects - Prospective.")
    ap.add_argument("--output", default=str(DATA_DIR / "prospective_scan.json"),
                    help="Output JSON path.")
    args = ap.parse_args()

    root = find_prospective_root()
    if not root:
        sys.exit("Could not locate '1.1 Projects - Prospective' under any known Dropbox root.")

    print(f"✓ Root: {root}")

    candidates = [c for c in root.iterdir() if c.is_dir()]
    if args.all:
        targets = candidates
    else:
        if not args.deals:
            sys.exit("Specify deal folder names as arguments, or use --all.")
        targets = []
        for q in args.deals:
            matches = [c for c in candidates if q.lower() in c.name.lower()]
            if not matches:
                print(f"⚠ No folder matched '{q}'", file=sys.stderr)
                continue
            targets.extend(matches)

    if not targets:
        sys.exit("No target folders found.")

    print(f"Scanning {len(targets)} folder(s):")
    for t in targets:
        print(f"  • {t.name}")
    print()

    results = {"deals": [scan_folder(t) for t in targets]}

    os.makedirs(os.path.dirname(args.output), exist_ok=True)
    with open(args.output, "w") as f:
        json.dump(results, f, default=str, indent=2)
    print(f"\n✓ Wrote {args.output}")
    total_bytes = os.path.getsize(args.output)
    print(f"  size: {total_bytes:,} bytes")
    return 0


if __name__ == "__main__":
    sys.exit(main())
